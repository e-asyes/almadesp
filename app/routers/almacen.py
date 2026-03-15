import io
import logging
import os
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Optional

import httpx
from bs4 import BeautifulSoup
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import get_current_user, TokenPayload
from app.database import get_db
from app.database_siscon import get_siscon_db
from app.models.manifiesto_bl import ManifiestoBL

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/almacen", tags=["almacen"])

ADUANA_URL = "https://isidora.aduana.cl/WebManifiestoMaritimo/Consultas/CON_BlsxMFTO.jsp?Action=Event"
ADUANA_ENABLED = os.environ.get("ADUANA_ENABLED", "true").lower() in ("true", "1", "yes")


# --- Response models ---

class BLDetail(BaseModel):
    n_bl: Optional[str] = None
    emisor: Optional[str] = None
    fecha_emision: Optional[str] = None
    fecha_aceptacion: Optional[str] = None
    fecha_embarque: Optional[str] = None
    almacen: Optional[str] = None
    puerto_embarque: Optional[str] = None
    puerto_desembarque: Optional[str] = None
    ultimo_transbordo: Optional[str] = None
    total_peso: Optional[str] = None


class ManifestHeader(BaseModel):
    nro_manifiesto: Optional[str] = None
    nave: Optional[str] = None
    sentido: Optional[str] = None
    fecha_arribo_zarpe: Optional[str] = None
    cia_naviera: Optional[str] = None
    fecha_emision_manifiesto: Optional[str] = None


class AlmacenManifest(BaseModel):
    header: ManifestHeader
    bls: list[BLDetail] = []


class AlmacenPartResult(BaseModel):
    bl_query: str
    manifests: list[AlmacenManifest] = []
    error: Optional[str] = None


class AlmacenLookupResponse(BaseModel):
    despacho: str
    numero_conocimiento: str
    parts: list[AlmacenPartResult]
    saved_count: int = 0


class ManifiestoBLResponse(BaseModel):
    id: int
    despacho: Optional[str] = None
    nro_manifiesto: str
    nave: Optional[str] = None
    sentido: Optional[str] = None
    fecha_arribo_zarpe: Optional[datetime] = None
    cia_naviera: Optional[str] = None
    fecha_emision_manifiesto: Optional[date] = None
    n_bl: Optional[str] = None
    almacen: Optional[str] = None
    fecha_aceptacion: Optional[date] = None
    puerto_desembarque: Optional[str] = None
    total_peso: Optional[Decimal] = None
    updated_at: Optional[datetime] = None
    almacen_real: Optional[str] = None
    usuario_actualizacion: Optional[str] = None
    fecha_actualizacion_manual: Optional[datetime] = None

    class Config:
        from_attributes = True


# --- Parsing helpers ---

def _cell_text(cell) -> str:
    return cell.get_text(strip=True) if cell else ""


def _find_label_value(soup: BeautifulSoup, label: str) -> Optional[str]:
    for td in soup.find_all("td"):
        if label in td.get_text():
            nxt = td.find_next_sibling("td")
            if nxt:
                val = _cell_text(nxt)
                return val if val else None
    return None


def _is_bl_table(table) -> bool:
    for td in table.find_all("td", class_="SimpleObjectTableCellTitle", recursive=True):
        if "BL" in td.get_text():
            return True
    return False


def _parse_manifests(html: str) -> list[AlmacenManifest]:
    soup = BeautifulSoup(html, "html.parser")
    manifests: list[AlmacenManifest] = []
    tables = soup.find_all("table")

    current_header: Optional[ManifestHeader] = None
    current_bls: list[BLDetail] = []

    for table in tables:
        table_class = " ".join(table.get("class", []))
        text = table.get_text()

        if "SimpleObjectTableCell" in table_class and "DATOS CONSIGNADOS" in text:
            if current_header is not None:
                manifests.append(AlmacenManifest(header=current_header, bls=current_bls))
                current_bls = []

            current_header = ManifestHeader(
                nro_manifiesto=_find_label_value(table, "Nro. Manifiesto"),
                nave=_find_label_value(table, "Nave"),
                sentido=_find_label_value(table, "Sentido"),
                fecha_arribo_zarpe=_find_label_value(table, "Fecha Arribo/Zarpe"),
                cia_naviera=_find_label_value(table, "Naviera"),
                fecha_emision_manifiesto=_find_label_value(table, "Fecha Emisi"),
            )

        elif "SimpleObjectTable" in table_class and _is_bl_table(table):
            rows = table.find_all("tr")
            for row in rows[1:]:
                cells = row.find_all("td")
                if len(cells) >= 10:
                    bl = BLDetail(
                        n_bl=_cell_text(cells[0]) or None,
                        emisor=_cell_text(cells[1]) or None,
                        fecha_emision=_cell_text(cells[2]) or None,
                        fecha_aceptacion=_cell_text(cells[3]) or None,
                        fecha_embarque=_cell_text(cells[4]) or None,
                        almacen=_cell_text(cells[5]) or None,
                        puerto_embarque=_cell_text(cells[6]) or None,
                        puerto_desembarque=_cell_text(cells[7]) or None,
                        ultimo_transbordo=_cell_text(cells[8]) or None,
                        total_peso=_cell_text(cells[9]) or None,
                    )
                    current_bls.append(bl)

    if current_header is not None:
        manifests.append(AlmacenManifest(header=current_header, bls=current_bls))

    return manifests


async def _query_bl(bl: str) -> AlmacenPartResult:
    form_data = {
        "EdNroGuia": bl,
        "EventSource": "cmdBuscar",
        "EventName": "buscarClick",
        "CON_ConsultaGralMFTOpageCode": "1",
        "totalManifiestos": "0",
    }

    try:
        async with httpx.AsyncClient(timeout=30.0, verify=False) as client:
            resp = await client.post(ADUANA_URL, data=form_data)
            resp.raise_for_status()
    except httpx.TimeoutException:
        logger.warning("Aduana timeout for BL %s", bl)
        return AlmacenPartResult(bl_query=bl, error="Aduana service timeout")
    except Exception as e:
        logger.warning("Aduana error for BL %s: %s: %s", bl, type(e).__name__, e)
        return AlmacenPartResult(bl_query=bl, error=f"Aduana service error: {type(e).__name__}: {e}")

    html = resp.content.decode("iso-8859-1")

    try:
        manifests = _parse_manifests(html)
    except Exception:
        logger.exception("Parse error for BL %s", bl)
        return AlmacenPartResult(bl_query=bl, error="Unable to parse response")

    if not manifests:
        return AlmacenPartResult(bl_query=bl, error="BL not found")

    return AlmacenPartResult(bl_query=bl, manifests=manifests)


def _parse_date(val: Optional[str]) -> Optional[date]:
    """Parse date string like '02-03-2026' or '24-02-2026' to date."""
    if not val:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(val: Optional[str]) -> Optional[datetime]:
    """Parse datetime string like '09/03/2026 15:17' to datetime."""
    if not val:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val.strip(), fmt)
        except ValueError:
            continue
    return None


def _parse_decimal(val: Optional[str]) -> Optional[Decimal]:
    """Parse numeric string to Decimal."""
    if not val:
        return None
    try:
        return Decimal(val.strip())
    except (InvalidOperation, ValueError):
        return None


async def _save_manifests_to_db(
    db: AsyncSession, despacho: str, manifests: list[AlmacenManifest]
) -> int:
    count = 0
    for manifest in manifests:
        header = manifest.header
        for bl in manifest.bls:
            if not bl.n_bl:
                continue

            # Check if a record with the same n_bl already exists
            existing = (
                await db.execute(
                    select(ManifiestoBL).where(ManifiestoBL.n_bl == bl.n_bl)
                )
            ).scalar_one_or_none()

            new_almacen = bl.almacen
            new_puerto = bl.puerto_desembarque
            new_nave = header.nave

            if existing:
                # Only update if almacen, puerto_desembarque or nave changed
                if (
                    existing.almacen == new_almacen
                    and existing.puerto_desembarque == new_puerto
                    and existing.nave == new_nave
                ):
                    continue

                existing.almacen = new_almacen
                existing.puerto_desembarque = new_puerto
                existing.nave = new_nave
                existing.despacho = despacho
                existing.nro_manifiesto = header.nro_manifiesto or ""
                existing.sentido = header.sentido
                existing.fecha_arribo_zarpe = _parse_datetime(header.fecha_arribo_zarpe)
                existing.cia_naviera = header.cia_naviera
                existing.fecha_emision_manifiesto = _parse_date(header.fecha_emision_manifiesto)
                existing.fecha_aceptacion = _parse_date(bl.fecha_aceptacion)
                existing.total_peso = _parse_decimal(bl.total_peso)
                count += 1
            else:
                record = ManifiestoBL(
                    despacho=despacho,
                    nro_manifiesto=header.nro_manifiesto or "",
                    nave=new_nave,
                    sentido=header.sentido,
                    fecha_arribo_zarpe=_parse_datetime(header.fecha_arribo_zarpe),
                    cia_naviera=header.cia_naviera,
                    fecha_emision_manifiesto=_parse_date(header.fecha_emision_manifiesto),
                    n_bl=bl.n_bl,
                    almacen=new_almacen,
                    fecha_aceptacion=_parse_date(bl.fecha_aceptacion),
                    puerto_desembarque=new_puerto,
                    total_peso=_parse_decimal(bl.total_peso),
                )
                db.add(record)
                count += 1
    return count


async def _get_bl_from_archimp(despacho: str, siscon_db: AsyncSession) -> Optional[str]:
    """Look up numero_conocimiento from archimp. Tries full despacho first, then without 502 prefix."""
    query = text(
        "SELECT numero_conocimiento FROM declaracion.archimp WHERE despacho = :despacho"
    )

    result = await siscon_db.execute(query, {"despacho": despacho})
    row = result.first()

    if not row and len(despacho) > 3:
        # Try without prefix (e.g. 5020529732 -> 0529732)
        short = despacho[3:]
        result = await siscon_db.execute(query, {"despacho": short})
        row = result.first()

    if row:
        return (row[0] or "").strip()
    return None


def _split_bl_queries(bl: str) -> list[str]:
    """Split BL on (H) or (N) notation markers."""
    upper = bl.upper()
    markers = re.findall(r'\([HN]\)', upper)
    if markers:
        parts_raw = re.split(r'(\([HN]\))', upper)
        queries = []
        for i, part in enumerate(parts_raw):
            part = part.strip()
            if not part:
                continue
            if re.match(r'\([HN]\)$', part):
                continue
            if i > 0 and re.match(r'\([HN]\)$', parts_raw[i - 1]):
                queries.append(parts_raw[i - 1] + part)
            else:
                queries.append(part)
        return queries
    return [bl]


# --- Endpoint ---

@router.get("/config")
async def get_config(current_user: TokenPayload = Depends(get_current_user)):
    """Return frontend configuration flags."""
    return {"aduana_enabled": ADUANA_ENABLED}


@router.get("/despacho/{despacho}", response_model=AlmacenLookupResponse)
async def lookup_almacen_by_despacho(
    despacho: str,
    db: AsyncSession = Depends(get_db),
    siscon_db: AsyncSession = Depends(get_siscon_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Look up Almacen info from Aduana by despacho number and save to manifiesto_bl."""
    if not ADUANA_ENABLED:
        raise HTTPException(status_code=503, detail="Consulta Aduana no disponible desde este servidor (IP no chilena)")
    despacho = despacho.strip()

    # 1. Get BL number from archimp
    numero_conocimiento = await _get_bl_from_archimp(despacho, siscon_db)
    if not numero_conocimiento:
        raise HTTPException(status_code=404, detail=f"Despacho {despacho} not found in archimp")

    # 2. Split BL and query aduana
    queries = _split_bl_queries(numero_conocimiento)

    parts: list[AlmacenPartResult] = []
    saved_count = 0

    for q in queries:
        result = await _query_bl(q)
        parts.append(result)

        if result.manifests:
            saved_count += await _save_manifests_to_db(db, despacho, result.manifests)

    return AlmacenLookupResponse(
        despacho=despacho,
        numero_conocimiento=numero_conocimiento,
        parts=parts,
        saved_count=saved_count,
    )


@router.get("/all", response_model=list[ManifiestoBLResponse])
async def list_all(
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """List all manifiesto_bl records."""
    result = await db.execute(
        select(ManifiestoBL).order_by(ManifiestoBL.id.desc())
    )
    return result.scalars().all()


@router.get("/almacenes-list")
async def list_almacenes(db: AsyncSession = Depends(get_db), current_user: TokenPayload = Depends(get_current_user)):
    """Return almacenes from the almacen_maestro table."""
    result = await db.execute(text("SELECT id, nombre, puerto FROM almacen_maestro ORDER BY nombre"))
    return [{"id": row[0], "nombre": row[1], "puerto": row[2]} for row in result.fetchall()]


class AlmacenMaestroCreate(BaseModel):
    nombre: str
    puerto: Optional[str] = None


@router.get("/almacenes")
async def list_almacenes_full(db: AsyncSession = Depends(get_db), current_user: TokenPayload = Depends(get_current_user)):
    """Return all almacenes from master table."""
    result = await db.execute(text("SELECT id, nombre, puerto FROM almacen_maestro ORDER BY nombre"))
    return [{"id": row[0], "nombre": row[1], "puerto": row[2]} for row in result.fetchall()]


@router.post("/almacenes")
async def create_almacen(
    payload: AlmacenMaestroCreate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Create a new almacen in the master table."""
    nombre = payload.nombre.strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre no puede estar vacio")
    existing = await db.execute(text("SELECT id FROM almacen_maestro WHERE nombre = :nombre"), {"nombre": nombre})
    if existing.first():
        raise HTTPException(status_code=409, detail=f"El almacen '{nombre}' ya existe")
    puerto = (payload.puerto or "").strip() or None
    result = await db.execute(
        text("INSERT INTO almacen_maestro (nombre, puerto) VALUES (:nombre, :puerto) RETURNING id, nombre, puerto"),
        {"nombre": nombre, "puerto": puerto}
    )
    row = result.first()
    return {"id": row[0], "nombre": row[1], "puerto": row[2]}


@router.put("/almacenes/{almacen_id}")
async def update_almacen(
    almacen_id: int,
    payload: AlmacenMaestroCreate,
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Update an almacen name."""
    nombre = payload.nombre.strip()
    if not nombre:
        raise HTTPException(status_code=400, detail="El nombre no puede estar vacio")
    existing = await db.execute(text("SELECT id FROM almacen_maestro WHERE id = :id"), {"id": almacen_id})
    if not existing.first():
        raise HTTPException(status_code=404, detail="Almacen no encontrado")
    dup = await db.execute(
        text("SELECT id FROM almacen_maestro WHERE nombre = :nombre AND id != :id"),
        {"nombre": nombre, "id": almacen_id}
    )
    if dup.first():
        raise HTTPException(status_code=409, detail=f"El almacen '{nombre}' ya existe")
    puerto = (payload.puerto or "").strip() or None
    result = await db.execute(
        text("UPDATE almacen_maestro SET nombre = :nombre, puerto = :puerto WHERE id = :id RETURNING id, nombre, puerto"),
        {"nombre": nombre, "puerto": puerto, "id": almacen_id}
    )
    row = result.first()
    return {"id": row[0], "nombre": row[1], "puerto": row[2]}


@router.delete("/almacenes/{almacen_id}")
async def delete_almacen(
    almacen_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Delete an almacen from the master table."""
    existing = await db.execute(text("SELECT id FROM almacen_maestro WHERE id = :id"), {"id": almacen_id})
    if not existing.first():
        raise HTTPException(status_code=404, detail="Almacen no encontrado")
    await db.execute(text("DELETE FROM almacen_maestro WHERE id = :id"), {"id": almacen_id})
    return {"ok": True}


@router.post("/almacenes/seed")
async def seed_almacenes(
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Populate almacen_maestro from existing distinct values in manifiesto_bl."""
    from sqlalchemy import union_all, literal_column
    q1 = select(ManifiestoBL.almacen.label("nombre")).where(ManifiestoBL.almacen.isnot(None))
    q2 = select(ManifiestoBL.almacen_real.label("nombre")).where(ManifiestoBL.almacen_real.isnot(None))
    combined = union_all(q1, q2).subquery()
    result = await db.execute(
        select(literal_column("nombre")).select_from(combined).distinct().order_by("nombre")
    )
    names = [row[0] for row in result.fetchall() if row[0] and row[0].strip()]
    added = 0
    for name in names:
        name = name.strip()
        existing = await db.execute(text("SELECT id FROM almacen_maestro WHERE nombre = :nombre"), {"nombre": name})
        if not existing.first():
            await db.execute(text("INSERT INTO almacen_maestro (nombre) VALUES (:nombre)"), {"nombre": name})
            added += 1
    return {"seeded": added, "total_distinct": len(names)}


class UpdateAlmacenRealRequest(BaseModel):
    almacen_real: str


class RegistroItem(BaseModel):
    despacho: str
    numero_conocimiento: Optional[str] = None
    puerto: Optional[str] = None
    eta: Optional[str] = None
    nombre_vehiculo: Optional[str] = None
    nombre_importador: Optional[str] = None
    status: str = "not_found"
    # Fields from manifiesto_bl (when found)
    id: Optional[int] = None
    n_bl: Optional[str] = None
    nave: Optional[str] = None
    almacen: Optional[str] = None
    puerto_desembarque: Optional[str] = None
    nro_manifiesto: Optional[str] = None
    cia_naviera: Optional[str] = None
    fecha_arribo_zarpe: Optional[datetime] = None
    fecha_aceptacion: Optional[date] = None
    total_peso: Optional[Decimal] = None
    updated_at: Optional[datetime] = None
    almacen_real: Optional[str] = None
    usuario_actualizacion: Optional[str] = None
    fecha_actualizacion_manual: Optional[datetime] = None


class RegistrosResponse(BaseModel):
    total: int = 0
    found: int = 0
    not_found: int = 0
    items: list[RegistroItem] = []


@router.get("/registros", response_model=RegistrosResponse)
async def list_registros(
    fecha_desde: str = Query("2026-03-13"),
    fecha_hasta: str = Query("2026-03-20"),
    puerto: str = Query(""),
    db: AsyncSession = Depends(get_db),
    siscon_db: AsyncSession = Depends(get_siscon_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """List all maritime despachos in date range, showing found and not-found."""
    port_filter = "AND TRIM(pto_desembarque) = :puerto" if puerto.strip() else ""
    query = text(f"""
        SELECT despacho, numero_conocimiento, TRIM(pto_desembarque) as puerto,
               fecha_arribo_estimado, nombre_vehiculo, nombre_importador
        FROM declaracion.archimp
        WHERE fecha_arribo_estimado BETWEEN :fecha_desde AND :fecha_hasta
          AND TRIM(cod_via_transporte) = '01'
          AND (__deleted IS NULL OR __deleted = 'false')
          {port_filter}
        ORDER BY fecha_arribo_estimado, despacho
    """)
    params: dict = {
        "fecha_desde": datetime.strptime(fecha_desde, "%Y-%m-%d").date(),
        "fecha_hasta": datetime.strptime(fecha_hasta, "%Y-%m-%d").date(),
    }
    if puerto.strip():
        params["puerto"] = puerto.strip().upper()

    archimp_rows = await siscon_db.execute(query, params)
    despachos = archimp_rows.fetchall()

    # Get all saved records from manifiesto_bl
    saved_result = await db.execute(select(ManifiestoBL))
    saved_records = saved_result.scalars().all()
    saved_by_despacho: dict[str, list] = {}
    for rec in saved_records:
        saved_by_despacho.setdefault(rec.despacho or "", []).append(rec)

    items: list[RegistroItem] = []
    found_count = 0
    not_found_count = 0

    for row in despachos:
        despacho = row[0]
        bl = (row[1] or "").strip()
        pto = row[2]
        eta = row[3]
        nombre_vehiculo = (row[4] or "").strip()
        nombre_importador = (row[5] or "").strip() or None
        eta_str = eta.strftime("%Y-%m-%d") if eta else None

        records = saved_by_despacho.get(despacho, [])
        if records:
            for rec in records:
                items.append(RegistroItem(
                    despacho=despacho, numero_conocimiento=bl, puerto=pto,
                    eta=eta_str, nombre_vehiculo=nombre_vehiculo,
                    nombre_importador=nombre_importador, status="found",
                    id=rec.id, n_bl=rec.n_bl, nave=rec.nave, almacen=rec.almacen,
                    puerto_desembarque=rec.puerto_desembarque,
                    nro_manifiesto=rec.nro_manifiesto, cia_naviera=rec.cia_naviera,
                    fecha_arribo_zarpe=rec.fecha_arribo_zarpe,
                    fecha_aceptacion=rec.fecha_aceptacion,
                    total_peso=rec.total_peso, updated_at=rec.updated_at,
                    almacen_real=rec.almacen_real,
                    usuario_actualizacion=rec.usuario_actualizacion,
                    fecha_actualizacion_manual=rec.fecha_actualizacion_manual,
                ))
            found_count += 1
        else:
            items.append(RegistroItem(
                despacho=despacho, numero_conocimiento=bl, puerto=pto,
                eta=eta_str, nombre_vehiculo=nombre_vehiculo,
                nombre_importador=nombre_importador, status="not_found",
            ))
            not_found_count += 1

    return RegistrosResponse(
        total=len(despachos), found=found_count,
        not_found=not_found_count, items=items,
    )


@router.put("/registros/{record_id}")
async def update_almacen_real(
    record_id: int,
    payload: UpdateAlmacenRealRequest,
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Update the real almacen value for a manifiesto_bl record. Requires auth."""
    result = await db.execute(
        select(ManifiestoBL).where(ManifiestoBL.id == record_id)
    )
    record = result.scalar_one_or_none()

    if not record:
        raise HTTPException(status_code=404, detail=f"Record {record_id} not found")

    record.almacen_real = payload.almacen_real.strip()
    record.usuario_actualizacion = current_user.email or current_user.preferred_username
    record.fecha_actualizacion_manual = datetime.now()

    await db.flush()
    await db.refresh(record)

    return {
        "id": record.id,
        "almacen": record.almacen,
        "almacen_real": record.almacen_real,
        "usuario_actualizacion": record.usuario_actualizacion,
        "fecha_actualizacion_manual": record.fecha_actualizacion_manual.isoformat() if record.fecha_actualizacion_manual else None,
    }


@router.get("/registros/excel")
async def download_registros_excel(
    fecha_desde: str = Query("2026-03-13"),
    fecha_hasta: str = Query("2026-03-20"),
    puerto: str = Query(""),
    db: AsyncSession = Depends(get_db),
    siscon_db: AsyncSession = Depends(get_siscon_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Download registros as Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Reuse the same logic from list_registros
    port_filter = "AND TRIM(pto_desembarque) = :puerto" if puerto.strip() else ""
    query = text(f"""
        SELECT despacho, numero_conocimiento, TRIM(pto_desembarque) as puerto,
               fecha_arribo_estimado, nombre_vehiculo, nombre_importador
        FROM declaracion.archimp
        WHERE fecha_arribo_estimado BETWEEN :fecha_desde AND :fecha_hasta
          AND TRIM(cod_via_transporte) = '01'
          AND (__deleted IS NULL OR __deleted = 'false')
          {port_filter}
        ORDER BY fecha_arribo_estimado, despacho
    """)
    params: dict = {
        "fecha_desde": datetime.strptime(fecha_desde, "%Y-%m-%d").date(),
        "fecha_hasta": datetime.strptime(fecha_hasta, "%Y-%m-%d").date(),
    }
    if puerto.strip():
        params["puerto"] = puerto.strip().upper()

    archimp_rows = await siscon_db.execute(query, params)
    despachos = archimp_rows.fetchall()

    saved_result = await db.execute(select(ManifiestoBL))
    saved_records = saved_result.scalars().all()
    saved_by_despacho: dict[str, list] = {}
    for rec in saved_records:
        saved_by_despacho.setdefault(rec.despacho or "", []).append(rec)

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    headers = [
        "Estado", "Despacho", "Importador", "Puerto", "ETA", "Nave/Vehiculo",
        "BL (Conocimiento)", "Nro BL (Aduana)", "Nave (Aduana)", "Almacen",
        "Almacen Real", "Puerto Destino", "Cia Naviera", "Nro Manifiesto",
        "Peso Total", "Actualizado", "Modificado Por", "Fecha Modificacion",
    ]

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row styles
    found_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    not_found_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    row_num = 2
    for row in despachos:
        despacho = row[0]
        bl = (row[1] or "").strip()
        pto = row[2]
        eta = row[3]
        nombre_vehiculo = (row[4] or "").strip()
        nombre_importador = (row[5] or "").strip()
        eta_str = eta.strftime("%Y-%m-%d") if eta else ""

        records = saved_by_despacho.get(despacho, [])
        if records:
            for rec in records:
                values = [
                    "ENCONTRADO", despacho, nombre_importador, pto, eta_str, nombre_vehiculo,
                    bl, rec.n_bl or "", rec.nave or "", rec.almacen or "",
                    rec.almacen_real or "",
                    rec.puerto_desembarque or "", rec.cia_naviera or "",
                    rec.nro_manifiesto or "",
                    float(rec.total_peso) if rec.total_peso else "",
                    rec.updated_at.strftime("%Y-%m-%d %H:%M") if rec.updated_at else "",
                    rec.usuario_actualizacion or "",
                    rec.fecha_actualizacion_manual.strftime("%Y-%m-%d %H:%M") if rec.fecha_actualizacion_manual else "",
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.fill = found_fill
                    cell.border = thin_border
                row_num += 1
        else:
            values = [
                "NO ENCONTRADO", despacho, nombre_importador, pto, eta_str, nombre_vehiculo,
                bl, "", "", "", "", "", "", "", "", "", "", "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = not_found_fill
                cell.border = thin_border
            row_num += 1

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, row_num):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 3, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    puerto_label = puerto.strip().upper() or "TODOS"
    filename = f"registros_{puerto_label}_{fecha_desde}_{fecha_hasta}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/bl/{n_bl}", response_model=list[ManifiestoBLResponse])
async def search_by_bl(
    n_bl: str,
    db: AsyncSession = Depends(get_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Search manifiesto_bl records by BL number (partial match)."""
    result = await db.execute(
        select(ManifiestoBL).where(ManifiestoBL.n_bl.ilike(f"%{n_bl.strip()}%"))
    )
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(status_code=404, detail=f"No records found for BL: {n_bl}")
    return rows


# --- Batch update by port and date range ---

class BLFound(BaseModel):
    n_bl: Optional[str] = None
    nave: Optional[str] = None
    almacen: Optional[str] = None
    puerto_desembarque: Optional[str] = None
    cia_naviera: Optional[str] = None
    fecha_arribo_zarpe: Optional[str] = None
    total_peso: Optional[str] = None
    saved: bool = False


class DespachoResult(BaseModel):
    despacho: str
    numero_conocimiento: str
    puerto: Optional[str] = None
    eta: Optional[str] = None
    nombre_vehiculo: Optional[str] = None
    status: str = "not_found"
    bls: list[BLFound] = []
    error: Optional[str] = None


class BatchUpdateResponse(BaseModel):
    puerto: str
    fecha_desde: str
    fecha_hasta: str
    total_despachos: int = 0
    total_found: int = 0
    total_not_found: int = 0
    total_saved: int = 0
    results: list[DespachoResult] = []


@router.get("/ports")
async def list_ports(
    siscon_db: AsyncSession = Depends(get_siscon_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """List available ports."""
    result = await siscon_db.execute(text("""
        SELECT DISTINCT TRIM(pto_desembarque) as puerto, COUNT(*) as total
        FROM declaracion.archimp
        WHERE TRIM(pto_desembarque) <> ''
          AND fecha_arribo_estimado >= '2025-01-01'
          AND TRIM(cod_via_transporte) = '01'
          AND (__deleted IS NULL OR __deleted = 'false')
        GROUP BY TRIM(pto_desembarque)
        ORDER BY total DESC
    """))
    return [{"puerto": row[0], "total": row[1]} for row in result.fetchall()]


@router.get("/batch-update", response_model=BatchUpdateResponse)
async def batch_update_by_port(
    puerto: str = Query("", description="Puerto de desembarque (vacio = todos)"),
    fecha_desde: str = Query("2026-03-13", description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: str = Query("2026-03-20", description="Fecha hasta (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
    siscon_db: AsyncSession = Depends(get_siscon_db),
    current_user: TokenPayload = Depends(get_current_user),
):
    """Batch update almacen data from Aduana for all despachos arriving at a port in a date range."""
    if not ADUANA_ENABLED:
        raise HTTPException(status_code=503, detail="Consulta Aduana no disponible desde este servidor (IP no chilena)")
    port_filter = "AND TRIM(pto_desembarque) = :puerto" if puerto.strip() else ""
    query = text(f"""
        SELECT despacho, numero_conocimiento, TRIM(pto_desembarque) as puerto,
               fecha_arribo_estimado, nombre_vehiculo
        FROM declaracion.archimp
        WHERE fecha_arribo_estimado BETWEEN :fecha_desde AND :fecha_hasta
          AND TRIM(numero_conocimiento) <> ''
          AND TRIM(cod_via_transporte) = '01'
          AND (__deleted IS NULL OR __deleted = 'false')
          {port_filter}
        ORDER BY fecha_arribo_estimado, despacho
    """)

    params: dict = {
        "fecha_desde": datetime.strptime(fecha_desde, "%Y-%m-%d").date(),
        "fecha_hasta": datetime.strptime(fecha_hasta, "%Y-%m-%d").date(),
    }
    if puerto.strip():
        params["puerto"] = puerto.strip().upper()

    result = await siscon_db.execute(query, params)
    rows = result.fetchall()

    results: list[DespachoResult] = []
    total_found = 0
    total_not_found = 0
    total_saved = 0

    for row in rows:
        despacho = row[0]
        numero_conocimiento = (row[1] or "").strip()
        pto = row[2]
        eta = row[3]
        nombre_vehiculo = (row[4] or "").strip()
        eta_str = eta.strftime("%Y-%m-%d") if eta else None

        if not numero_conocimiento:
            results.append(DespachoResult(
                despacho=despacho, numero_conocimiento="", puerto=pto,
                eta=eta_str, nombre_vehiculo=nombre_vehiculo,
                status="not_found", error="No BL number",
            ))
            total_not_found += 1
            continue

        queries = _split_bl_queries(numero_conocimiento)
        found_bls: list[BLFound] = []
        despacho_found = False

        for q in queries:
            bl_result = await _query_bl(q)
            if bl_result.manifests:
                despacho_found = True
                saved = await _save_manifests_to_db(db, despacho, bl_result.manifests)
                total_saved += saved
                for m in bl_result.manifests:
                    for bl in m.bls:
                        found_bls.append(BLFound(
                            n_bl=bl.n_bl,
                            nave=m.header.nave,
                            almacen=bl.almacen,
                            puerto_desembarque=bl.puerto_desembarque,
                            cia_naviera=m.header.cia_naviera,
                            fecha_arribo_zarpe=m.header.fecha_arribo_zarpe,
                            total_peso=bl.total_peso,
                            saved=saved > 0,
                        ))

        if despacho_found:
            total_found += 1
            results.append(DespachoResult(
                despacho=despacho, numero_conocimiento=numero_conocimiento,
                puerto=pto, eta=eta_str, nombre_vehiculo=nombre_vehiculo,
                status="found", bls=found_bls,
            ))
        else:
            total_not_found += 1
            results.append(DespachoResult(
                despacho=despacho, numero_conocimiento=numero_conocimiento,
                puerto=pto, eta=eta_str, nombre_vehiculo=nombre_vehiculo,
                status="not_found", error="BL not found in Aduana",
            ))

    return BatchUpdateResponse(
        puerto=puerto.strip().upper() or "TODOS",
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        total_despachos=len(rows),
        total_found=total_found,
        total_not_found=total_not_found,
        total_saved=total_saved,
        results=results,
    )
